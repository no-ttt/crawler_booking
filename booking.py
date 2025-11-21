from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import quote
import time
import pandas as pd
import re
from datetime import datetime, timedelta
# --- 從 format.py 整合進來的函式庫 ---
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


def setup_driver(headless=False):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1400,1000")
    driver = webdriver.Chrome(options=options)
    return driver

def scrape_hotel_data(driver, wait, hotel_url, checkin_date):
    print(f"\n--- 開始處理飯店連結 ---\n🔗 {hotel_url}")
    
    hotel_title = "N/A"
    
    try:
        driver.get(hotel_url)
        # 使用您指定的、非常精確的 CSS Selector 來定位飯店標題
        # 提醒：此 Selector 較為脆弱，若網站結構變動可能失效
        css_selector = "#wrap-hotelpage-top > div:nth-child(3) > div > div.ecb8d66605.f228f8d929.daadf70613 > h2"
        title_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, css_selector)))
        hotel_title = title_element.text.strip().split('\n')[0] # 取第一行以防抓到多餘文字
        print(f"🏨 飯店名稱: {hotel_title}")
    except Exception as e:
        print(f"❌ 抓取飯店標題失敗: {e}")
        return []

    all_room_data = []

    try:
        hprt_table = wait.until(
            EC.presence_of_element_located((By.ID, "hprt-table"))
        )
        print("✅ 找到房型價格表 (hprt-table)。")

        room_rows = hprt_table.find_elements(By.CSS_SELECTOR, "tbody tr")
        print(f"🔎 找到 {len(room_rows)} 種房型，開始擷取價格...")

        current_room_name = "N/A"
        current_occupancy = "N/A"

        for row in room_rows:
            try:
                # 嘗試抓取房型名稱。如果某個 tr 沒有房型名稱，代表它跟上一個是同一個房型
                room_name_element = row.find_element(By.CSS_SELECTOR, ".hprt-roomtype-link")
                current_room_name = room_name_element.text.strip()

                # 抓取人數 (通常跟房型名稱在同一個 tr)
                occupancy_element = row.find_element(By.CSS_SELECTOR, ".hprt-occupancy-occupancy-info .bui-u-sr-only")
                occupancy_text = occupancy_element.get_attribute("textContent").strip()

                # 從文字中只取出數字
                occupancy_match = re.search(r'\d+', occupancy_text)
                if occupancy_match:
                    current_occupancy = occupancy_match.group(0)
                
                # 抓取價格
                price_text = row.find_element(By.CSS_SELECTOR, ".bui-price-display__value").text.strip()
                # 只保留數字部分 (移除貨幣符號、逗號等)
                price = re.sub(r'\D', '', price_text)
                
                room_info = {
                    "網址": hotel_url,
                    "飯店名": hotel_title,
                    "入住日期": checkin_date,
                    "房型": current_room_name,
                    "人數": current_occupancy,
                    "價格": price,
                }
                all_room_data.append(room_info)
                print(f"✅ 成功擷取: {room_info}")

            except Exception as e:
                # 如果在上面 try 區塊出錯，可能是同房型但不同價格的列 (沒有房型名稱)
                # 嘗試只抓取價格，並沿用上一個房型的名稱和人數
                try:
                    price_text = row.find_element(By.CSS_SELECTOR, ".bui-price-display__value").text.strip()
                    # 只保留數字部分 (移除貨幣符號、逗號等)
                    price = re.sub(r'\D', '', price_text)
                    room_info = {
                        "網址": hotel_url,
                        "飯店名": hotel_title,
                        "入住日期": checkin_date,
                        "房型": f"{current_room_name}",
                        "人數": current_occupancy,
                        "價格": price
                    }
                    all_room_data.append(room_info)
                    print(f"✅ 成功擷取: {room_info}")
                except Exception:
                    # 如果連價格都抓不到，代表這可能只是個分隔行，直接跳過
                    continue
    except Exception as e:
        print("⚠️ 飯店頁抓不到價格，可能頁面結構更新或需要滑動", e)
    
    return all_room_data

def format_and_save_excel(df):
    if df.empty:
        print("🤷‍♂️ 沒有資料可供格式化。")
        return

    max_rooms = int(df.groupby(["飯店名", "入住日期"]).size().max())

    # 建立新 Excel
    wb = Workbook()
    ws = wb.active

    # 樣式設定
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # ====== 第 1 列：大標題 ======
    first_row = ["入住日期", "網址", "飯店名"] + ["房價"] + [""] * (max_rooms - 1)
    ws.append(first_row)

    # 合併「房價」標題
    if max_rooms > 1:
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3 + max_rooms)

    # 樣式設定
    for col in range(1, 4 + max_rooms):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = align_center
        cell.border = border

    # ====== 寫入資料 ======
    start_row = 2
    for (hotel, date), group in df.groupby(["飯店名", "入住日期"]):
        url = group["網址"].iloc[0] if "網址" in group.columns else ""

        rooms = list(group["房型"])
        prices = list(group["價格"])

        # 補齊房型數
        while len(rooms) < max_rooms:
            rooms.append("")
            prices.append("")

        # 上排（房型）
        row_room_names = [date, url, hotel] + rooms[:max_rooms]
        # 下排（價格）
        row_prices = row_room_names[:3] + prices[:max_rooms] # 直接複製上排資訊，更穩健

        ws.append(row_room_names)
        ws.append(row_prices)

        # 合併「入住日期」「網址」「飯店名」
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + 1, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row + 1, end_column=3)

        # 樣式設定
        for col in range(1, 4 + max_rooms):
            top_cell = ws.cell(row=start_row, column=col)
            bot_cell = ws.cell(row=start_row + 1, column=col)

            top_cell.border = border
            bot_cell.border = border
            top_cell.alignment = align_center
            bot_cell.alignment = align_center

            if col > 3:  # 房型行灰底
                top_cell.fill = gray_fill
                bot_cell.fill = white_fill

        start_row += 2

    # 自動調整欄寬
    for i, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        max_length = 0
        for cell in col_cells:
            if cell.value:
                # 考慮中文寬度，稍微加權
                cell_len = 0
                for char in str(cell.value):
                    cell_len += 2 if '\u4e00' <= char <= '\u9fff' else 1
                max_length = max(max_length, cell_len)
        ws.column_dimensions[col_letter].width = max_length + 3

    # 儲存輸出
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"{timestamp_str}.xlsx"
    wb.save(output_path)

    print(f"\n🎉 全部完成！格式化報告已存入 {output_path}")


if __name__ == "__main__":
    # --- 動態設定查詢日期 ---
    # 1. checkin_date 設定為今天日期後一週
    checkin_datetime = datetime.now() + timedelta(days=7)
    checkin_date = checkin_datetime.strftime("%Y-%m-%d")
    # 2. checkout_date 設定為 checkin_date 再加一天
    checkout_datetime = checkin_datetime + timedelta(days=1)
    checkout_date = checkout_datetime.strftime("%Y-%m-%d")

    hotel_urls = [
        f"https://www.booking.com/hotel/tw/dou-dian-inn.zh-tw.html?checkin={checkin_date}&checkout={checkout_date}&group_adults=2&no_rooms=1",
        f"https://www.booking.com/hotel/tw/fu-ye-wen-quan-xiu-xian-hui-guan.zh-tw.html?checkin={checkin_date}&checkout={checkout_date}&group_adults=2&no_rooms=1"
    ]

    print(f"\n{'='*20} 開始爬取 (標準模式) {'='*20}")
    print(f"執行時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    master_data_list = []
    
    # --- 初始設定：開啟瀏覽器 ---
    driver = setup_driver(headless=False)
    wait = WebDriverWait(driver, 20)

    for url_index, url in enumerate(hotel_urls):
        current_hotel_data = []
        max_browser_restarts_for_url = 3 # 對每個 URL，如果抓取失敗，最多重啟瀏覽器嘗試的次數
        
        for restart_attempt in range(max_browser_restarts_for_url):
            print(f"\n--- 處理飯店連結 {url_index + 1}/{len(hotel_urls)} (瀏覽器重啟嘗試 {restart_attempt + 1}/{max_browser_restarts_for_url}) ---")
            
            # 如果不是第一次嘗試 (restart_attempt > 0)，表示上次抓取失敗，需要重啟瀏覽器
            if restart_attempt > 0:
                print("   抓取失敗，關閉並重開瀏覽器...")
                driver.quit()
                driver = setup_driver(headless=False)
                wait = WebDriverWait(driver, 20)

            # 嘗試抓取資料
            current_hotel_data = scrape_hotel_data(driver, wait, url, checkin_date)
            
            if current_hotel_data:
                print(f"✅ 飯店連結 {url_index + 1} 抓取成功！")
                master_data_list.extend(current_hotel_data)
                break # 成功抓取，跳出當前 URL 的重啟嘗試迴圈，處理下一個 URL
            else:
                print(f"⚠️ 飯店連結 {url_index + 1} 抓取失敗。")
                if restart_attempt < max_browser_restarts_for_url - 1:
                    print("   等待 3 秒後重試此飯店連結...")
                    time.sleep(3)
                    # 迴圈會繼續，並在下一次嘗試前重啟瀏覽器
                else:
                    print(f"❌ 連續 {max_browser_restarts_for_url} 次嘗試失敗，跳過此飯店連結。")
                    break # 跳出當前 URL 的重啟嘗試迴圈，處理下一個 URL
        
        time.sleep(3) # 每次處理完一個飯店連結後休息一下，避免被封鎖

    # 最後，關閉所有瀏覽器實例（如果還有開啟的）
    print("\n關閉最後一個瀏覽器實例。")
    driver.quit()

    # --- 將爬取結果直接傳遞給格式化函式 ---
    if master_data_list:
        df = pd.DataFrame(master_data_list)
        format_and_save_excel(df)
    else:
        print("\n🤷‍♂️ 沒有抓取到任何房型資料。")
    
